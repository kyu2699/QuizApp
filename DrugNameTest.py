import openpyxl
import os
import random
import tkinter as tk
from tkinter import messagebox
import sys
import subprocess


class DrugQuiz:
    def __init__(self, sheet_name, sheet, root, start_button, excel_file):
        self.sheet_name = sheet_name
        self.sheet = sheet
        self.score = 0
        self.total = 0
        self.root = root
        self.start_button = start_button
        self.row_numbers = list(range(2, self.sheet.max_row + 1))
        random.shuffle(self.row_numbers)
        self.excel_file = excel_file

        self.create_ui_components()

        self.current_row = 0
        self.display_question()

        self.start_button.config(state="disabled")

    def create_ui_components(self):
        """Create all the UI components for the quiz."""
        self.header_frame = tk.Frame(self.root, bg="#E6E6FA")
        self.header_frame.pack(fill="x", padx=10, pady=10)
        
        # Main content
        self.question_label = tk.Label(self.root, text="Ready to start the quiz!", font=("Arial", 16))
        self.question_label.pack(pady=20)

        self.answer_entry = tk.Entry(self.root, font=("Arial", 14))
        self.answer_entry.pack(pady=10)

        self.submit_button = tk.Button(self.root, text="Submit Answer", font=("Arial", 14), command=self.submit_answer)
        self.submit_button.pack(pady=10)

        self.result_label = tk.Label(self.root, text="", font=("Arial", 12))
        self.result_label.pack(pady=20)

        self.score_label = tk.Label(self.root, text=f"Score: {self.score}", font=("Arial", 14))
        self.score_label.pack(anchor="nw", padx=10, pady=10)

        self.history_frame = tk.Frame(self.root, bg="black")
        self.history_frame.pack(pady=10, fill="x")

        self.history_label = tk.Label(self.history_frame, text="History:", font=("Arial", 14), fg="white", bg="black")
        self.history_label.pack()

        self.history_text = tk.Label(self.history_frame, text="", font=("Arial", 12), justify="left", fg="white", bg="black")
        self.history_text.pack()

        # Add copyright text
        self.copyright_label = tk.Label(self.root, text="By yours truly, Kyle Yu for Kyra Lynn", font=("Arial", 8), fg="gray", bg="#E6E6FA")
        self.copyright_label.pack(side="right", padx=10, pady=10)

        # Home Button
        self.home_button = tk.Button(self.root, text="Home", font=("Arial", 14), command=self.go_home)
        self.home_button.place(x=10, y=self.root.winfo_height() - 50, anchor="sw")

        # Exit Button
        self.exit_button = tk.Button(self.root, text="Exit", font=("Arial", 14), command=self.exit_quiz)
        self.exit_button.place(x=100, y=self.root.winfo_height() - 50, anchor="sw")

        # Edit Answer Key Button
        self.edit_button = tk.Button(self.root, 
                                    text="Edit Answer Key", 
                                    font=("Arial", 14), 
                                    command=self.edit_answer_key, 
                                    fg="RED",
                                    relief="solid",
                                    bd=4,
                                    highlightthickness=4,
                                    highlightbackground="red",
                                    highlightcolor="red")
        self.edit_button.place(x=200, y=self.root.winfo_height() - 50, anchor="sw")


        self.root.bind("<Return>", self.submit_answer_from_key)

    def display_question(self):
        """Display the next question or complete the quiz if finished."""
        if self.current_row < len(self.row_numbers):
            next_row = self.row_numbers[self.current_row]
            drug_name = self.sheet.cell(row=next_row, column=1).value
            self.question_label.config(text=f"What is the Brand/Generic Name for '{drug_name}'?")
        else:
            self.show_final_score()

    def show_final_score(self):
        """Display the final score and handle post-quiz actions."""
        self.result_label.config(text=f"Final Score: {self.score}/{self.total}")
        self.submit_button.config(state="disabled")
        self.answer_entry.config(state="disabled")
        self.quiz_complete_popup()

    def quiz_complete_popup(self):
        """Show a popup when the quiz is complete and ask if the user wants to retry."""
        response = messagebox.askquestion(
            "Quiz Complete",
            f"You have completed the quiz for this section.\nYour score: {self.score}/{self.total}\nWould you like to retry with another sheet or exit?",
        )
        if response == 'yes':
            self.root.destroy() 
            main()
        else:
            self.exit_quiz()

    def submit_answer(self, event=None):
        """Handle the submission of an answer and check if it's correct."""
        user_input = self.answer_entry.get().strip().lower()
        if not user_input:
            self.result_label.config(text="Please enter an answer!", fg="orange")
            return

        if self.current_row < len(self.row_numbers):
            row = self.row_numbers[self.current_row]
            column_a_value = self.sheet.cell(row=row, column=1).value
            correct_answers = []
            for col in range(2, self.sheet.max_column + 1):
                cell_value = self.sheet.cell(row=row, column=col).value
                if cell_value:
                    correct_answers.append(str(cell_value).strip().lower())

            correct_answer = user_input in correct_answers

            self.update_result_and_score(correct_answer, correct_answers)
            self.display_question()

        else:
            self.show_final_score()

    def update_result_and_score(self, correct_answer, correct_answers):
        """Update the result label, score, history, and move to the next question."""
        if correct_answer:
            self.result_label.config(text="Correct!", fg="green")
            self.score += 1
        else:
            self.result_label.config(text="Wrong!", fg="red")

        self.total += 1
        self.current_row += 1
        self.answer_entry.delete(0, tk.END)

        self.score_label.config(text=f"Score: {self.score}/{self.total}")
        previous_question = self.sheet.cell(row=self.row_numbers[self.current_row - 1], column=1).value
        if previous_question is not None:
            correct_answer_display = " or ".join(correct_answers) if correct_answers else "No valid answer"
            history_content = f"Q: {previous_question}\nCorrect Answer(s): {correct_answer_display}\n"
            self.history_text.config(text=history_content)
        else:
            self.history_text.config(text="")

    def go_home(self):
        """Return to sheet selection window."""
        self.root.destroy() 
        main()

    def exit_quiz(self):
        """Exit the quiz application."""
        self.root.quit()

    def submit_answer_from_key(self, event):
        """Allow submission of answer using the Enter key."""
        self.submit_answer()

    def edit_answer_key(self):
        """Open the Excel sheet for editing."""
        try:
            # Get the absolute path of the Excel file
            if getattr(sys, 'frozen', False):
                base_path = os.path.dirname(sys.executable)
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
            file_path = os.path.join(base_path, self.excel_file)
            if sys.platform == "win32":
                os.startfile(file_path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", file_path])
            else:
                subprocess.Popen(["xdg-open", file_path])

            messagebox.showinfo("Info", "You can now edit the answer key. Make sure to save and close the file before returning to the quiz.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while trying to open the Excel file: {e}")


class ExcelQuiz:
    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook = self.load_excel(file_name)
        self.sheets = self.workbook.sheetnames

    def load_excel(self, file_name='drugnames.xlsx'):
        """Load the Excel file."""
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.executable)
        else: 
            base_path = os.path.dirname(os.path.abspath(__file__))

        file_path = os.path.join(base_path, file_name)

        if not os.path.exists(file_path):
            messagebox.showerror("Error", f"File '{file_name}' not found.")
            sys.exit(1)
        return openpyxl.load_workbook(file_path)

    def prompt_user_for_sheet(self, root, start_button):
        """Prompt the user to select a sheet for the quiz."""
        def on_sheet_selected():
            sheet_choice = sheet_listbox.curselection()
            if sheet_choice:
                sheet_name = self.sheets[sheet_choice[0]]
                sheet = self.workbook[sheet_name]
                DrugQuiz(sheet_name, sheet, root, start_button, self.file_name)
                sheet_selection_window.destroy()

        sheet_selection_window = tk.Toplevel(root)
        sheet_selection_window.title("Select Sheet")
        sheet_selection_window.geometry("400x300")
        sheet_listbox = tk.Listbox(sheet_selection_window, height=len(self.sheets), font=("Arial", 14))
        for sheet in self.sheets:
            sheet_listbox.insert(tk.END, sheet)
        sheet_listbox.pack(pady=20)

        select_button = tk.Button(sheet_selection_window, text="Select", font=("Arial", 14), command=on_sheet_selected)
        select_button.pack(pady=10)


def main():
    """Main entry point to start the quiz application."""
    quiz_app = ExcelQuiz(file_name="drugnames.xlsx")

    root = tk.Tk()
    root.title("Drug Quiz")
    root.state("zoomed")
    root.geometry("800x600")
    root.config(bg="#E6E6FA")

    welcome_label = tk.Label(root, text="Welcome to the Drug Naming Quiz!", font=("Arial", 24), pady=20, bg="#E6E6FA")
    welcome_label.pack()

    description_label = tk.Label(root, text="This quiz will test your knowledge of drug name vocabulary.\nYou will be asked to provide the Brand/Generic name for various drugs.\nLet's see how many you can get right!", font=("Arial", 16), justify="center", padx=20, bg="#E6E6FA")
    description_label.pack(pady=20)

    instructions_label = tk.Label(root, text="Instructions:\n1. Choose a sheet with drug names.\n2. Answer each question with the correct Brand/Generic name.\n3. You can exit anytime by clicking 'Exit'.\n4. Your score is displayed as you progress.", font=("Arial", 14), justify="left", padx=20, bg="#E6E6FA")
    instructions_label.pack(pady=20)

    start_button = tk.Button(root, text="Start Quiz", font=("Arial", 20), command=lambda: quiz_app.prompt_user_for_sheet(root, start_button), bg="#E6E6FA")
    start_button.pack(pady=50)

    root.mainloop()


if __name__ == "__main__":
    main()
